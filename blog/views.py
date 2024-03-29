import os
from django import forms
from django.conf import settings
from django.shortcuts import redirect
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.views.generic import TemplateView
from .models import Post 
from .forms import PostForm
from django.core.files.storage import FileSystemStorage


from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render

import openpyxl
from openpyxl import Workbook, load_workbook

import json


def post_test(request):
    form = PostForm()
    return render(request, 'blog/teston.html', {'form': form})
def post_list(request):
    posts = Post.objects.filter(published_date__lte=timezone.now()).order_by('published_date')
    return render(request, 'blog/post_list.html', {'posts': posts})
def post_detail(request, pk):
    post = get_object_or_404(Post, pk=pk)
    return render(request, 'blog/post_detail.html', {'post': post})
def post_new(request):
    form = PostForm()
    return render(request, 'blog/post_edit.html', {'form': form})
def post_new(request):
    if request.method == "POST":
        form = PostForm(request.POST)
        if form.is_valid():
            post = form.save(commit=False)
            post.author = request.user
            post.published_date = timezone.now()
            post.save()
            return redirect('post_detail', pk=post.pk)
    else:
        form = PostForm()
    return render(request, 'blog/post_edit.html', {'form': form})
def post_edit(request, pk):
    post = get_object_or_404(Post, pk=pk)
    if request.method == "POST":
        form = PostForm(request.POST, instance=post)
        if form.is_valid():
            post = form.save(commit=False)
            post.author = request.user
            post.published_date = timezone.now()
            post.save()
            return redirect('post_detail', pk=post.pk)
    else:
        form = PostForm(instance=post)
    return render(request, 'blog/post_edit.html', {'form': form})
def excel(request):
    return render(request,'blog/excel.html')

def handle_uploaded_file(f):
    with open('some/file/name.txt', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)



    

#def columns(request):
    #lastfile = File.objects.last()
    #filepath = lastfile.filepath
    #filename= lastfile.name
    #form= FileForm(request.POST or None, request.FILES or None)
    #if form.is_valid():
    #    form.save()
    #context = {'filepath': filepath,
     #'form': form,
      #'filename':filename
      #}
    #return render(request, 'blog/excel.html', context)


def post_columns(request):

    if request.method=='POST':
        # print(request)
        uploaded_file= request.FILES['file']
        fs = FileSystemStorage()
        fs.save(uploaded_file.name, uploaded_file)
       
       # Utiliser OpenPyXL pour manipuler le fichier xls
        wb = load_workbook(filename=uploaded_file, read_only=True)
        n=0
        sheets = wb.sheetnames
        ws = wb[sheets[n]]
        #for row in ws.rows:
          #  print(row[0])

        #colnames = []
        #col_indices = {n for n, cell in enumerate(ws.rows[0]) if cell.value in colnames}
        data =[[cell.value for cell in row] for row in ws.rows]
        y = json.dumps({"items":data[0]})

    return HttpResponse(y, content_type="application/json")





    #myfile = request.FILES['upload']
    #fs = FileSystemStorage()
    #filename = fs.save(myfile.name, myfile)
    #uploaded_file_url = fs.url(filename)
    #render(request, 'core/simple_upload.html', {
     #   'uploaded_file_url': uploaded_file_url
    #})
    #return {}
    
        #book = load_workbook()
        #sheet = book.active
        #data =[[cell.value for cell in row] for row in sheet.rows]
            #data[0]